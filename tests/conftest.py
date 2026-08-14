"""Fixtures shared by more than one test module.

The client records and the template builders are used by both the module that owns
them and the invoice-generation tests, so they live here rather than being copied.
"""

import openpyxl
import pytest

import time_tracker_template


CLIENT_HOURLY = {
    "company": "FakeCo",
    "contact": "Fake Client",
    "phone": "(123) 456-8900",
    "addr1": "Fake addr1",
    "addr2": "Fake City, FS 00000",
    "rate_hr": 100,
    "rate_day": 800,
}

CLIENT_RETAINER = {
    **CLIENT_HOURLY,
    "company": "Retainer Corp",
    "retainer_hrs": 20,
    "retainer_rate": 1500,
}

# The variables every template must declare; see REQUIRED_TEMPLATE_VARIABLES.
REQUIRED_ROWS = (
    ("invoice_num", None),
    ("invoice_date", None),
    ("cust_company", None),
    ("cust_contact", None),
    ("print_area_ul", "A1"),
    ("print_area_lr", "C38"),
)


def build_template_workbook(rows=REQUIRED_ROWS, sheet_name=None):
    """Build an in-memory template whose variables sheet declares ``rows``."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name if sheet_name is not None else time_tracker_template.VARIABLES_SHEET_NAME
    sheet.cell(row=1, column=1, value="Variable")
    sheet.cell(row=1, column=2, value="Value")

    for index, (name, value) in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=name)
        if value is not None:
            sheet.cell(row=index, column=2, value=value)

    return workbook


def save_template(tmp_path, **kwargs) -> str:
    """Write a built template to ``tmp_path`` and return its path."""
    path = tmp_path / "template.xlsx"
    build_template_workbook(**kwargs).save(path)
    return str(path)


@pytest.fixture
def template_workbook():
    """The in-memory template builder, for tests that never touch the disk."""
    return build_template_workbook


@pytest.fixture
def saved_template(tmp_path):
    """A template written to disk, for the functions that take a path."""
    def _saved(**kwargs) -> str:
        return save_template(tmp_path, **kwargs)

    return _saved