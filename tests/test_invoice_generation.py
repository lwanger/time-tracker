"""Tests for invoice generation in time_tracker_invoice.

Covers the amounts, the values written to a template, and make_invoice end to end
against the shipped template. Excel COM is mocked throughout - the PDF export is the
one step that needs a real Excel.
"""

import datetime
import sys
import types
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

import time_tracker_invoice
import time_tracker_template
import time_tracker_templates
from conftest import CLIENT_HOURLY, CLIENT_RETAINER, save_template


@pytest.fixture(autouse=True)
def on_windows(monkeypatch):
    """Run every test in this module as if on Windows, where the PDF export exists.

    Without this the make_invoice tests would assert an export that never happens on a
    Linux or macOS CI runner. The two tests that describe the off-Windows behaviour set
    ``sys.platform`` again themselves; the later monkeypatch wins.
    """
    monkeypatch.setattr(sys, "platform", "win32")


def test_previous_month_steps_back_within_year():
    assert time_tracker_invoice.previous_month(datetime.date(2026, 7, 26)) == (2026, 6)


def test_previous_month_january_rolls_back_a_year():
    assert time_tracker_invoice.previous_month(datetime.date(2026, 1, 15)) == (2025, 12)


# --------------------------------------------------------------------------- #
# export_pdf (Excel COM is mocked)
# --------------------------------------------------------------------------- #
class _FakeComError(Exception):
    """Stands in for pywintypes.com_error.

    The real exception cannot be raised - or even imported - off Windows, which is what
    stopped this module being collected there at all.
    """


def _com_error(description: str) -> _FakeComError:
    """Build a com_error shaped like the ones Excel raises over COM."""
    return _FakeComError(
        -2147352567,
        "Exception occurred.",
        (0, "Microsoft Excel", description, "xlmain11.chm", 0, -2146827284),
        None,
    )


@pytest.fixture
def fake_excel(monkeypatch):
    """Fake ``pywintypes`` / ``win32com.client`` modules, and the Excel they hand back.

    ``export_pdf`` imports both from inside the function, since pywin32 installs on
    Windows alone, so there is no module attribute left to patch. What the local import
    reads is ``sys.modules``, so that is what these tests replace - which also makes
    ``except com_error`` catch :class:`_FakeComError` rather than the real thing.

    Returns a namespace of the ``DispatchEx`` callable and the Excel object it returns.
    """
    excel = MagicMock()
    dispatch = MagicMock(return_value=excel)

    pywintypes = types.ModuleType("pywintypes")
    pywintypes.com_error = _FakeComError

    win32com = types.ModuleType("win32com")
    win32com_client = types.ModuleType("win32com.client")
    win32com_client.DispatchEx = dispatch
    win32com.client = win32com_client

    monkeypatch.setitem(sys.modules, "pywintypes", pywintypes)
    monkeypatch.setitem(sys.modules, "win32com", win32com)
    monkeypatch.setitem(sys.modules, "win32com.client", win32com_client)

    return types.SimpleNamespace(excel=excel, dispatch=dispatch)


def test_export_pdf_drives_excel_com(fake_excel):
    time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")

    fake_excel.dispatch.assert_called_once_with("Excel.Application")
    worksheet = fake_excel.excel.Workbooks.Open.return_value.Worksheets.return_value
    assert worksheet.PageSetup.PrintArea == "A1:H40"
    worksheet.ExportAsFixedFormat.assert_called_once()
    fake_excel.excel.Quit.assert_called_once()


def test_export_pdf_exports_only_the_invoice_sheet(fake_excel):
    """Exporting the workbook would append the Variables sheet to the client's PDF."""
    time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")

    workbook = fake_excel.excel.Workbooks.Open.return_value
    workbook.Worksheets.assert_called_once_with(1)
    workbook.Worksheets.return_value.ExportAsFixedFormat.assert_called_once()
    workbook.ExportAsFixedFormat.assert_not_called()


def test_export_pdf_opens_an_absolute_path_without_updating_links(fake_excel):
    """Excel resolves relative paths against its own cwd, and must not chase links."""
    time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")

    opened_path, open_kwargs = fake_excel.excel.Workbooks.Open.call_args
    assert Path(opened_path[0]).is_absolute()
    assert Path(opened_path[0]).name == "book.xlsx"
    assert open_kwargs == {"UpdateLinks": 0}
    assert fake_excel.excel.DisplayAlerts is False


def test_export_pdf_raises_invoice_export_error_not_com_error(fake_excel):
    """No win32 exception may escape: a caller off Windows cannot even name the type."""
    fake_excel.excel.Workbooks.Open.side_effect = _com_error("Open method of Workbooks class failed")

    with pytest.raises(time_tracker_invoice.InvoiceExportError, match="Open method"):
        time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")


def test_export_pdf_quits_excel_when_opening_fails(fake_excel):
    """DispatchEx starts a dedicated Excel process; a failure must not orphan it."""
    fake_excel.excel.Workbooks.Open.side_effect = _com_error("Open method of Workbooks class failed")

    with pytest.raises(time_tracker_invoice.InvoiceExportError):
        time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")

    fake_excel.excel.Quit.assert_called_once()


def test_export_pdf_closes_workbook_and_quits_when_export_fails(fake_excel):
    workbook = fake_excel.excel.Workbooks.Open.return_value
    workbook.Worksheets.return_value.ExportAsFixedFormat.side_effect = _com_error("export failed")

    with pytest.raises(time_tracker_invoice.InvoiceExportError):
        time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")

    workbook.Close.assert_called_once_with(SaveChanges=False)
    fake_excel.excel.Quit.assert_called_once()


def test_export_pdf_quits_excel_even_when_closing_fails(fake_excel):
    workbook = fake_excel.excel.Workbooks.Open.return_value
    workbook.Close.side_effect = _com_error("close failed")

    # A com_error raised while shutting Excel down is converted too, so the caller sees
    # one exception type whatever failed.
    with pytest.raises(time_tracker_invoice.InvoiceExportError, match="close failed"):
        time_tracker_invoice.export_pdf("book.xlsx", ("A1", "H40"), "book.pdf")

    fake_excel.excel.Quit.assert_called_once()


# --------------------------------------------------------------------------- #
# build_invoice_variables / build_invoice_line_items
# --------------------------------------------------------------------------- #
JULY_2026 = datetime.date(2026, 7, 26)


def test_build_invoice_variables_hourly_client():
    values = time_tracker_invoice.build_invoice_variables(CLIENT_HOURLY, "TEST", 150, 10, today=JULY_2026)

    assert values["invoice_total"] == 1000.0
    assert values["invoice_amt1"] == 1000.0
    assert values["invoice_hours"] == 10
    assert values["invoice_period"] == "June 2026"
    assert (values["invoice_year"], values["invoice_month"]) == (2026, 6)
    assert values["cust_code"] == "TEST"
    assert values["invoice_rate_per_hour"] == 100
    assert values["invoice_rate_per_day"] == 800
    # An hourly client has no retainer figures to display.
    assert values["invoice_retainer_amount"] is None
    assert values["invoice_overage_hours"] is None
    assert values["invoice_overage_amount"] is None


def test_build_invoice_variables_leaves_phone_unlabelled():
    # The template concatenates its own "Phone: " prefix, so the value must be bare.
    values = time_tracker_invoice.build_invoice_variables(CLIENT_HOURLY, "TEST", 1, 1, today=JULY_2026)

    assert values["cust_phone"] == CLIENT_HOURLY["phone"]


def test_build_invoice_variables_retainer_with_overage():
    values = time_tracker_invoice.build_invoice_variables(CLIENT_RETAINER, "RET", 151, 30, today=JULY_2026)

    assert values["invoice_total"] == 2500.0
    assert values["invoice_retainer_amount"] == 1500
    assert values["invoice_overage_hours"] == 10
    assert values["invoice_overage_amount"] == 1000.0
    assert values["invoice_amt2"] == 1500
    assert values["invoice_amt3"] == 1000.0
    assert values["cust_retainer_hrs"] == 20


def test_build_invoice_variables_retainer_without_overage():
    values = time_tracker_invoice.build_invoice_variables(CLIENT_RETAINER, "RET", 152, 15, today=JULY_2026)

    assert values["invoice_total"] == 1500.0
    assert values["invoice_overage_hours"] is None
    assert "invoice_desc3" not in values
    assert "invoice_amt3" not in values


def test_build_invoice_line_items_indents_retainer_sublines():
    items = time_tracker_invoice.build_invoice_line_items(CLIENT_RETAINER, 30, 2500.0, "June 2026", indent="   ")

    assert items["invoice_desc2"].startswith("   Retained services")
    assert items["invoice_desc3"].startswith("   Additional hours")


# Hours reach these lines as logged minutes / 60, so they carry binary-float noise:
# 1232 minutes / 60 is 20.533333333333335, and 20.533333333333335 - 20 is 0.5333333333333350.
NOISY_HOURS = 1232 / 60


def test_build_invoice_line_items_rounds_hourly_hours_to_one_decimal():
    items = time_tracker_invoice.build_invoice_line_items(
        CLIENT_HOURLY, NOISY_HOURS, 2053.33, "July 2026", indent="",
    )

    assert items["invoice_desc1"] == (
        "Hourly consulting services - July 2026 (20.5 hours @ $100.00/hr)"
    )


def test_build_invoice_line_items_rounds_retainer_hours_to_one_decimal():
    items = time_tracker_invoice.build_invoice_line_items(
        CLIENT_RETAINER, NOISY_HOURS, 1553.33, "July 2026", indent="",
    )

    assert items["invoice_desc2"] == (
        "Retained services - 1st 20.0 hours @ fixed rate of $1500.00/mo"
    )
    assert items["invoice_desc3"] == "Additional hours - 0.5 hours @ $100.00/hr"


def test_build_invoice_line_items_never_shows_a_long_float():
    """Guards every description line, not just the ones asserted verbatim above."""
    for client in (CLIENT_HOURLY, CLIENT_RETAINER):
        items = time_tracker_invoice.build_invoice_line_items(
            client, NOISY_HOURS, 2053.33, "July 2026", indent="",
        )
        descriptions = [value for name, value in items.items() if name.startswith("invoice_desc")]
        assert descriptions
        for description in descriptions:
            assert "3333" not in description, description


# --------------------------------------------------------------------------- #
# make_invoice (real openpyxl against the shipped template; export_pdf mocked)
# --------------------------------------------------------------------------- #
class _JanuaryDate(datetime.date):
    """date subclass whose today() is in January, to exercise the year rollover."""

    @classmethod
    def today(cls):
        return datetime.date(2026, 1, 15)


def _shipped_template() -> str:
    return str(time_tracker_templates.template_path())


def test_make_invoice_writes_values_into_the_variables_sheet(tmp_path):
    with patch("time_tracker_invoice.export_pdf") as mock_pdf:
        time_tracker_invoice.make_invoice(
            _shipped_template(), {"TEST": CLIENT_HOURLY},
            inv_num=150, client_name="TEST", inv_hrs=10, inv_save_dir=str(tmp_path),
        )

    saved = list(tmp_path.glob("Invoice 150*.xlsx"))
    assert len(saved) == 1
    mock_pdf.assert_called_once()

    workbook = openpyxl.load_workbook(saved[0])
    rows = time_tracker_invoice.read_template_variable_rows(workbook)
    sheet = workbook[time_tracker_template.VARIABLES_SHEET_NAME]
    assert sheet.cell(row=rows["invoice_num"], column=2).value == 150
    assert sheet.cell(row=rows["invoice_total"], column=2).value == 1000.0
    assert sheet.cell(row=rows["cust_company"], column=2).value == "FakeCo"


def test_make_invoice_takes_print_area_from_the_template(tmp_path):
    # Read the expected bounds from the template itself: they are the template's to
    # choose, so hardcoding them here would just re-assert a copy of the data.
    template = openpyxl.load_workbook(_shipped_template())
    template_rows = time_tracker_invoice.read_template_variable_rows(template)
    expected = tuple(
        time_tracker_invoice.read_template_setting(template, template_rows, name)
        for name in time_tracker_template.PRINT_AREA_VARIABLES
    )

    with patch("time_tracker_invoice.export_pdf") as mock_pdf:
        time_tracker_invoice.make_invoice(
            _shipped_template(), {"TEST": CLIENT_HOURLY},
            inv_num=151, client_name="TEST", inv_hrs=1, inv_save_dir=str(tmp_path),
        )

    assert mock_pdf.call_args.args[1] == expected
    assert all(bound for bound in expected), "template must declare both print-area bounds"


def test_shipped_template_indent_is_whitespace_only():
    # The shipped template once held a literal backtick before the five spaces - a
    # mix-up with Excel's text-force prefix, which is an apostrophe and is stored as a
    # style flag rather than as data. `indent` is prefixed to the retainer sub-line
    # descriptions, so any visible character in it prints on the client's invoice.
    workbook = openpyxl.load_workbook(_shipped_template())
    rows = time_tracker_invoice.read_template_variable_rows(workbook)
    indent = time_tracker_invoice.read_template_setting(workbook, rows, "indent", default="")

    assert not indent.strip(), f"indent must be whitespace only, got {indent!r}"


def test_make_invoice_indents_retainer_sublines_from_the_shipped_template(tmp_path):
    """The shipped template's own `indent` reaching a retainer invoice's descriptions.

    The other shipped-template tests bill an hourly client, whose line items never use
    `indent`, so nothing exercised the template's value end to end.
    """
    with patch("time_tracker_invoice.export_pdf"):
        time_tracker_invoice.make_invoice(
            _shipped_template(), {"RET": CLIENT_RETAINER},
            inv_num=152, client_name="RET", inv_hrs=25, inv_save_dir=str(tmp_path),
        )

    workbook = openpyxl.load_workbook(next(tmp_path.glob("Invoice 152*.xlsx")))
    rows = time_tracker_invoice.read_template_variable_rows(workbook)
    sheet = workbook[time_tracker_template.VARIABLES_SHEET_NAME]

    for name, expected in (
        ("invoice_desc2", "Retained services - 1st 20.0 hours @ fixed rate of $1500.00/mo"),
        ("invoice_desc3", "Additional hours - 5.0 hours @ $100.00/hr"),
    ):
        description = sheet.cell(row=rows[name], column=2).value
        # lstrip() strips whitespace only, so equality here also proves the indent
        # carries no visible character into what the client sees.
        assert description.lstrip() == expected
        assert description.startswith(" "), f"{name} is not indented: {description!r}"


def test_make_invoice_without_rate_raises_before_opening_the_template(tmp_path):
    """A rate-less client must fail before any file work.

    The rate is not otherwise read until the invoice-log append, which runs after both
    the .xlsx and the .pdf exist - a bare KeyError there orphans them.
    """
    rate_less = {key: value for key, value in CLIENT_HOURLY.items() if key != "rate_hr"}

    with (
        patch("time_tracker_invoice.xl.load_workbook") as mock_load,
        patch("time_tracker_invoice.export_pdf") as mock_pdf,
        pytest.raises(time_tracker_invoice.ClientRateError, match="rate_hr"),
    ):
        time_tracker_invoice.make_invoice(
            _shipped_template(), {"PRO": rate_less},
            inv_num=153, client_name="PRO", inv_hrs=5, inv_save_dir=str(tmp_path),
        )

    mock_load.assert_not_called()
    mock_pdf.assert_not_called()
    assert not list(tmp_path.glob("Invoice*"))


def test_make_invoice_without_variables_sheet_writes_nothing(tmp_path):
    template = save_template(tmp_path, sheet_name="Simple Invoice")

    with (
        patch("time_tracker_invoice.export_pdf") as mock_pdf,
        pytest.raises(time_tracker_template.TemplateError),
    ):
        time_tracker_invoice.make_invoice(template, {"TEST": CLIENT_HOURLY}, inv_num=1,
                                   client_name="TEST", inv_hrs=1, inv_save_dir=str(tmp_path))

    # The check runs before anything is written, so there is nothing to clean up.
    assert not list(tmp_path.glob("Invoice*"))
    mock_pdf.assert_not_called()


def test_make_invoice_missing_required_variable_writes_nothing(tmp_path):
    template = save_template(
        tmp_path, rows=(("invoice_num", None), ("print_area_ul", "A1"), ("print_area_lr", "C1")),
    )

    with pytest.raises(time_tracker_template.TemplateError, match="missing required variables"):
        time_tracker_invoice.make_invoice(template, {"TEST": CLIENT_HOURLY}, inv_num=1,
                                    client_name="TEST", inv_hrs=1, inv_save_dir=str(tmp_path))

    assert not list(tmp_path.glob("Invoice*"))


def test_make_invoice_returns_the_workbook_it_wrote(tmp_path):
    with patch("time_tracker_invoice.export_pdf"):
        invoice_path = time_tracker_invoice.make_invoice(
            _shipped_template(), {"TEST": CLIENT_HOURLY},
            inv_num=152, client_name="TEST", inv_hrs=4, inv_save_dir=str(tmp_path),
        )

    assert invoice_path == next(tmp_path.glob("Invoice 152*.xlsx"))


def test_make_invoice_removes_the_workbook_when_pdf_export_fails(tmp_path):
    """The .xlsx is saved before the PDF is exported, so a failure must not leave it."""
    export_failure = time_tracker_invoice.InvoiceExportError("Open method of Workbooks class failed")

    with (
        patch("time_tracker_invoice.export_pdf", side_effect=export_failure),
        pytest.raises(time_tracker_invoice.InvoiceExportError, match="could not export invoice 153"),
    ):
        time_tracker_invoice.make_invoice(
            _shipped_template(), {"TEST": CLIENT_HOURLY},
            inv_num=153, client_name="TEST", inv_hrs=4, inv_save_dir=str(tmp_path),
        )

    assert not list(tmp_path.glob("Invoice*"))


# --------------------------------------------------------------------------- #
# make_invoice off Windows: the .xlsx is the invoice, the PDF a rendering of it
# --------------------------------------------------------------------------- #
def test_pdf_export_supported_follows_the_platform(monkeypatch):
    assert time_tracker_invoice.pdf_export_supported()  # the autouse fixture says win32

    monkeypatch.setattr(sys, "platform", "linux")
    assert not time_tracker_invoice.pdf_export_supported()


def test_make_invoice_off_windows_writes_the_xlsx_and_skips_the_export(tmp_path, monkeypatch):
    monkeypatch.setattr(sys, "platform", "linux")

    with patch("time_tracker_invoice.export_pdf") as mock_pdf:
        invoice_path = time_tracker_invoice.make_invoice(
            _shipped_template(), {"TEST": CLIENT_HOURLY},
            inv_num=154, client_name="TEST", inv_hrs=4, inv_save_dir=str(tmp_path),
        )

    mock_pdf.assert_not_called()
    assert invoice_path == next(tmp_path.glob("Invoice 154*.xlsx"))
    assert not list(tmp_path.glob("*.pdf"))


def test_make_invoice_off_windows_needs_no_pywin32(tmp_path, monkeypatch):
    """The test that actually proves the dependency is optional.

    A Windows CI job installs the real pywin32, so nothing else here would notice an
    import creeping back to module scope. Binding a module to None in ``sys.modules``
    makes importing it raise ImportError, which is what a Linux runner would see.
    """
    monkeypatch.setattr(sys, "platform", "linux")
    monkeypatch.setitem(sys.modules, "pywintypes", None)
    monkeypatch.setitem(sys.modules, "win32com", None)
    monkeypatch.setitem(sys.modules, "win32com.client", None)

    time_tracker_invoice.make_invoice(
        _shipped_template(), {"TEST": CLIENT_HOURLY},
        inv_num=155, client_name="TEST", inv_hrs=4, inv_save_dir=str(tmp_path),
    )

    assert list(tmp_path.glob("Invoice 155*.xlsx"))


def test_make_invoice_january_rolls_back_to_previous_december(tmp_path):
    with (
        patch("time_tracker_invoice.export_pdf"),
        patch("time_tracker_invoice.datetime.date", _JanuaryDate),
    ):
        time_tracker_invoice.make_invoice(
            _shipped_template(), {"TEST": CLIENT_HOURLY},
            inv_num=160, client_name="TEST", inv_hrs=5, inv_save_dir=str(tmp_path),
        )

    assert list(tmp_path.glob("Invoice 160 - 2025-12*.xlsx"))


def test_make_invoice_prompts_for_client_and_hours(tmp_path):
    with (
        patch("time_tracker_invoice.export_pdf"),
        patch("time_tracker_invoice.ci.get_string", autospec=True, return_value="TEST") as mock_get_str,
        patch("time_tracker_invoice.ci.get_float", autospec=True, return_value=12.5) as mock_get_float,
    ):
        time_tracker_invoice.make_invoice(_shipped_template(), {"TEST": CLIENT_HOURLY},
                                   inv_num=153, inv_save_dir=str(tmp_path))

    mock_get_str.assert_called_once()
    mock_get_float.assert_called_once()
    assert list(tmp_path.glob("Invoice 153*.xlsx"))


# --------------------------------------------------------------------------- #
# compute_invoice_total
# --------------------------------------------------------------------------- #
def test_compute_invoice_total_hourly_client():
    assert time_tracker_invoice.compute_invoice_total(CLIENT_HOURLY, 10) == 1000.0


def test_compute_invoice_total_retainer_with_additional_hours():
    # 20 hrs @ 1500 flat + 10 additional hrs @ 100/hr = 2500.
    assert time_tracker_invoice.compute_invoice_total(CLIENT_RETAINER, 30) == 2500.0


def test_compute_invoice_total_retainer_no_additional_hours():
    # Under the retained hours -> just the flat retainer fee.
    assert time_tracker_invoice.compute_invoice_total(CLIENT_RETAINER, 15) == 1500.0
