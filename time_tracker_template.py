"""The Invoice Template's Variables Sheet: reading, validating and filling it in.

A template declares the variables it uses on its own worksheet rather than against an
external cell map, so the invoice sheet can reference them by formula and Excel keeps
those references correct when cells move. See
docs/adr/0003-self-describing-invoice-template.md.

Leonard Wanger, 2026
"""

import re
from typing import Any

import openpyxl as xl


# Invoice-template variables sheet. See docs/adr/0003-self-describing-invoice-template.md:
# a template declares its own variables here instead of an external cell map, so the
# invoice sheet can reference them by formula and Excel keeps the references correct
# when cells move.
VARIABLES_SHEET_NAME = "Variables"
VARIABLE_NAME_COL = 1
VARIABLE_VALUE_COL = 2

# A row declares a variable when its name cell is a bare lowercase identifier. This
# skips the header row ("Variable") and the "[ ... ]" section headers without needing
# to know which rows they occupy, so template authors can rearrange the sheet freely.
VARIABLE_NAME_RE = re.compile(r"^[a-z][a-z0-9_]*$")

# Variables whose absence makes a generated invoice wrong or unprintable, so it is an
# error rather than a template that simply does not use them. Every other name is
# optional by design - Time Tracker offers a vocabulary and each template draws from it.
REQUIRED_TEMPLATE_VARIABLES = (
    "invoice_num",
    "invoice_date",
    "cust_company",
    "cust_contact",
    "print_area_ul",
    "print_area_lr",
)

# Template settings read by Time Tracker rather than written to.
PRINT_AREA_VARIABLES = ("print_area_ul", "print_area_lr")


class TemplateError(Exception):
    """An invoice template lacks its variables sheet or a required variable."""


def read_template_variable_rows(workbook: xl.Workbook) -> dict[str, int]:
    """Map every variable declared on a template's variables sheet to its row.

    Args:
        workbook: An open invoice-template workbook.

    Returns:
        Mapping of variable name to its 1-based row on the variables sheet.

    Raises:
        TemplateError: If the workbook has no variables sheet.
    """
    if VARIABLES_SHEET_NAME not in workbook.sheetnames:
        raise TemplateError(
            f"template has no '{VARIABLES_SHEET_NAME}' worksheet "
            f"(sheets present: {', '.join(workbook.sheetnames)})"
        )

    sheet = workbook[VARIABLES_SHEET_NAME]
    variable_rows: dict[str, int] = {}

    for row_index in range(1, sheet.max_row + 1):
        name = sheet.cell(row=row_index, column=VARIABLE_NAME_COL).value
        if isinstance(name, str) and VARIABLE_NAME_RE.match(name.strip()):
            variable_rows[name.strip()] = row_index

    return variable_rows


def read_template_setting(
    workbook: xl.Workbook, variable_rows: dict[str, int], name: str, default: Any = None
) -> Any:
    """Read a template-configuration value (e.g. ``indent``) from the variables sheet.

    Args:
        workbook: An open invoice-template workbook.
        variable_rows: Mapping from :func:`read_template_variable_rows`.
        name: Variable name to read.
        default: Returned when the variable is undeclared or its value cell is empty.

    Returns:
        The cell's value, or ``default``.
    """
    row = variable_rows.get(name)
    if row is None:
        return default

    value = workbook[VARIABLES_SHEET_NAME].cell(row=row, column=VARIABLE_VALUE_COL).value
    return default if value is None else value


def validate_template_variables(
    workbook: xl.Workbook, variable_rows: dict[str, int], template_file: str
) -> None:
    """Check a template declares everything Time Tracker cannot generate an invoice without.

    Called before any file is written, so a bad template fails loudly rather than
    silently producing an invoice with blank fields.

    Args:
        workbook: An open invoice-template workbook.
        variable_rows: Mapping from :func:`read_template_variable_rows`.
        template_file: Path to the template, for the error message.

    Raises:
        TemplateError: If a required variable is undeclared, or a print-area
            setting is declared but has no value.
    """
    missing = [name for name in REQUIRED_TEMPLATE_VARIABLES if name not in variable_rows]
    if missing:
        raise TemplateError(
            f"template is missing required variables: {', '.join(missing)}. "
            f"Add them to column A of the '{VARIABLES_SHEET_NAME}' sheet in {template_file}"
        )

    blank = [
        name
        for name in PRINT_AREA_VARIABLES
        if not read_template_setting(workbook, variable_rows, name)
    ]
    if blank:
        raise TemplateError(
            f"template has no value for: {', '.join(blank)}. "
            f"Set them in column B of the '{VARIABLES_SHEET_NAME}' sheet in {template_file}"
        )

    check_template_has_no_external_links(workbook, template_file)


def check_template_has_no_external_links(workbook: xl.Workbook, template_file: str) -> None:
    """Reject a template that links to another workbook.

    A template copied from another workbook inherits an Excel external link, so
    formulas read ``[1]Variables!B22`` - the *source* workbook's sheet - instead of
    the template's own. Two things then go wrong, and the second is silent:

    * openpyxl cannot round-trip the link. Excel writes it as an ``rId1`` relative
      relationship plus an ``xxl21:alternateUrls`` ``rId2``; openpyxl keeps only
      ``rId2`` while still emitting ``<externalBook r:id="rId1">``. Excel rejects the
      dangling relationship, which surfaces from COM as "Open method of Workbooks
      class failed" once :func:`export_pdf` tries to open the generated invoice.
    * The linked cells resolve against the wrong workbook, so the invoice shows stale
      values rather than the client being billed.

    Args:
        workbook: An open invoice-template workbook.
        template_file: Path to the template, for the error message.

    Raises:
        TemplateError: If the template has any external workbook link.
    """
    # openpyxl exposes external links only through this private attribute; the public
    # API has no accessor for them. Stable across openpyxl 3.1.x.
    external_links = getattr(workbook, "_external_links", [])
    if not external_links:
        return

    targets = ", ".join(
        str(link.file_link.Target) for link in external_links if link.file_link is not None
    )
    raise TemplateError(
        f"template links to another workbook ({targets or 'unknown target'}), so its "
        f"'{VARIABLES_SHEET_NAME}' formulas read that workbook instead of its own and "
        f"Excel cannot open the generated invoice. Repair it with "
        f"'python tools/repair_template_external_links.py \"{template_file}\"', or in "
        f"Excel replace '[1]{VARIABLES_SHEET_NAME}!' with '{VARIABLES_SHEET_NAME}!' in "
        f"the invoice sheet's formulas"
    )


def write_template_variables(
    workbook: xl.Workbook, variable_rows: dict[str, int], values: dict[str, Any]
) -> None:
    """Write invoice values into a template's variables sheet, addressed by name.

    Names the template does not declare are skipped rather than treated as errors: a
    template references only the variables it needs. ``None`` values are skipped too,
    leaving the cell empty so the invoice sheet's ``IF(...<>"")`` guards apply.

    Args:
        workbook: An open invoice-template workbook.
        variable_rows: Mapping from :func:`read_template_variable_rows`.
        values: Variable name to value.
    """
    sheet = workbook[VARIABLES_SHEET_NAME]

    for name, value in values.items():
        row = variable_rows.get(name)
        if row is not None and value is not None:
            sheet.cell(row=row, column=VARIABLE_VALUE_COL, value=value)
